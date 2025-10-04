"""
Report generators for PDF and Excel export functionality
"""
import io
import csv
from datetime import datetime, timedelta
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from django.db.models import Count, Avg, Q
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from core.models import Region, UserProfile
from alerts.models import Alert, AlertDelivery
from drought_data.models import DroughtRiskAssessment, WeatherData, NDVIData
from farmers.models import FarmerProfile
from ussd.models import USSDSession, USSDUser


class DroughtReportGenerator:
    """Generate comprehensive drought monitoring reports"""
    
    def __init__(self, start_date=None, end_date=None):
        self.end_date = end_date or timezone.now().date()
        self.start_date = start_date or (self.end_date - timedelta(days=30))
    
    def generate_csv_report(self, report_type='summary'):
        """Generate CSV report"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="drought_report_{report_type}_{self.start_date}_to_{self.end_date}.csv"'
        
        writer = csv.writer(response)
        
        if report_type == 'summary':
            self._write_summary_csv(writer)
        elif report_type == 'assessments':
            self._write_assessments_csv(writer)
        elif report_type == 'alerts':
            self._write_alerts_csv(writer)
        elif report_type == 'weather':
            self._write_weather_csv(writer)
        
        return response
    
    def generate_excel_report(self, report_type='full'):
        """Generate comprehensive Excel report"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Add summary sheet
        self._add_summary_sheet(wb)
        
        if report_type == 'full':
            # Add detailed sheets
            self._add_assessments_sheet(wb)
            self._add_alerts_sheet(wb)
            self._add_weather_sheet(wb)
            self._add_farmers_sheet(wb)
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="drought_report_{report_type}_{self.start_date}_to_{self.end_date}.xlsx"'
        
        wb.save(response)
        return response
    
    def generate_pdf_report(self, report_type='summary'):
        """Generate PDF report"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1  # Center
        )
        story.append(Paragraph("Drought Monitoring Report", title_style))
        story.append(Paragraph(f"Period: {self.start_date} to {self.end_date}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        if report_type == 'summary':
            self._add_summary_pdf_content(story, styles)
        elif report_type == 'detailed':
            self._add_detailed_pdf_content(story, styles)
        
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="drought_report_{report_type}_{self.start_date}_to_{self.end_date}.pdf"'
        response.write(buffer.getvalue())
        
        return response
    
    def _write_summary_csv(self, writer):
        """Write summary data to CSV"""
        writer.writerow(['Drought Monitoring Summary Report'])
        writer.writerow(['Period', f'{self.start_date} to {self.end_date}'])
        writer.writerow(['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        writer.writerow([])
        
        # Regional summary
        writer.writerow(['Regional Summary'])
        writer.writerow(['Region', 'Latest Risk Score', 'Risk Level', 'Alerts Sent', 'Last Assessment'])
        
        regions = Region.objects.filter(region_type='county')
        for region in regions:
            latest_assessment = DroughtRiskAssessment.objects.filter(
                region=region,
                assessment_date__range=[self.start_date, self.end_date]
            ).order_by('-assessment_date').first()
            
            alerts_count = Alert.objects.filter(
                region=region,
                created_at__date__range=[self.start_date, self.end_date]
            ).count()
            
            if latest_assessment:
                writer.writerow([
                    region.name,
                    f"{latest_assessment.risk_score:.1f}",
                    latest_assessment.get_risk_level_display(),
                    alerts_count,
                    latest_assessment.assessment_date
                ])
            else:
                writer.writerow([region.name, 'No data', 'No data', alerts_count, 'No assessment'])
    
    def _write_assessments_csv(self, writer):
        """Write drought assessments to CSV"""
        writer.writerow([
            'Assessment ID', 'Region', 'Date', 'Risk Score', 'Risk Level',
            'NDVI Score', 'Soil Moisture Score', 'Weather Score',
            'Predicted 7-day Risk', 'Predicted 30-day Risk', 'Confidence'
        ])
        
        assessments = DroughtRiskAssessment.objects.filter(
            assessment_date__range=[self.start_date, self.end_date]
        ).select_related('region').order_by('-assessment_date')
        
        for assessment in assessments:
            writer.writerow([
                assessment.id,
                assessment.region.name,
                assessment.assessment_date,
                f"{assessment.risk_score:.1f}",
                assessment.get_risk_level_display(),
                f"{assessment.ndvi_component_score:.1f}",
                f"{assessment.soil_moisture_component_score:.1f}",
                f"{assessment.weather_component_score:.1f}",
                f"{assessment.predicted_risk_7_days:.1f}" if assessment.predicted_risk_7_days else 'N/A',
                f"{assessment.predicted_risk_30_days:.1f}" if assessment.predicted_risk_30_days else 'N/A',
                f"{assessment.confidence_score:.2f}"
            ])
    
    def _write_alerts_csv(self, writer):
        """Write alerts data to CSV"""
        writer.writerow([
            'Alert ID', 'Region', 'Title', 'Type', 'Severity', 'Priority',
            'Status', 'Created Date', 'Sent Date', 'Recipients', 'Successful', 'Failed'
        ])
        
        alerts = Alert.objects.filter(
            created_at__date__range=[self.start_date, self.end_date]
        ).select_related('region').order_by('-created_at')
        
        for alert in alerts:
            writer.writerow([
                alert.alert_id,
                alert.region.name,
                alert.title,
                alert.get_template_display() if hasattr(alert, 'get_template_display') else 'N/A',
                alert.severity_level if hasattr(alert, 'severity_level') else 'N/A',
                alert.priority,
                alert.status,
                alert.created_at.date(),
                alert.sent_at.date() if alert.sent_at else 'Not sent',
                alert.total_recipients,
                alert.successfully_sent,
                alert.failed_sends
            ])
    
    def _write_weather_csv(self, writer):
        """Write weather data to CSV"""
        writer.writerow([
            'Region', 'Date', 'Temperature', 'Humidity', 'Rainfall',
            'Wind Speed', 'Solar Radiation', 'Barometric Pressure'
        ])
        
        weather_data = WeatherData.objects.filter(
            timestamp__date__range=[self.start_date, self.end_date]
        ).select_related('region').order_by('-timestamp')
        
        for data in weather_data:
            writer.writerow([
                data.region.name,
                data.timestamp.date(),
                f"{data.temperature:.1f}" if data.temperature else 'N/A',
                f"{data.humidity:.1f}" if data.humidity else 'N/A',
                f"{data.rainfall:.1f}" if data.rainfall else 'N/A',
                f"{data.wind_speed:.1f}" if data.wind_speed else 'N/A',
                f"{data.solar_radiation:.1f}" if data.solar_radiation else 'N/A',
                f"{data.barometric_pressure:.1f}" if data.barometric_pressure else 'N/A'
            ])
    
    def _add_summary_sheet(self, wb):
        """Add summary sheet to Excel workbook"""
        ws = wb.create_sheet(title="Summary")
        
        # Headers
        ws['A1'] = "Drought Monitoring Summary Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Period: {self.start_date} to {self.end_date}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Regional data
        ws['A5'] = "Regional Summary"
        ws['A5'].font = Font(bold=True)
        
        headers = ['Region', 'Latest Risk Score', 'Risk Level', 'Alerts Sent', 'Last Assessment']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Data
        regions = Region.objects.filter(region_type='county')
        row = 7
        for region in regions:
            latest_assessment = DroughtRiskAssessment.objects.filter(
                region=region,
                assessment_date__range=[self.start_date, self.end_date]
            ).order_by('-assessment_date').first()
            
            alerts_count = Alert.objects.filter(
                region=region,
                created_at__date__range=[self.start_date, self.end_date]
            ).count()
            
            ws.cell(row=row, column=1, value=region.name)
            if latest_assessment:
                ws.cell(row=row, column=2, value=f"{latest_assessment.risk_score:.1f}")
                ws.cell(row=row, column=3, value=latest_assessment.get_risk_level_display())
                ws.cell(row=row, column=5, value=str(latest_assessment.assessment_date))
            else:
                ws.cell(row=row, column=2, value="No data")
                ws.cell(row=row, column=3, value="No data")
                ws.cell(row=row, column=5, value="No assessment")
            
            ws.cell(row=row, column=4, value=alerts_count)
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_assessments_sheet(self, wb):
        """Add drought assessments sheet"""
        ws = wb.create_sheet(title="Assessments")
        
        headers = [
            'Assessment ID', 'Region', 'Date', 'Risk Score', 'Risk Level',
            'NDVI Score', 'Soil Moisture Score', 'Weather Score',
            'Predicted 7-day Risk', 'Predicted 30-day Risk', 'Confidence'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        assessments = DroughtRiskAssessment.objects.filter(
            assessment_date__range=[self.start_date, self.end_date]
        ).select_related('region').order_by('-assessment_date')
        
        for row, assessment in enumerate(assessments, 2):
            ws.cell(row=row, column=1, value=assessment.id)
            ws.cell(row=row, column=2, value=assessment.region.name)
            ws.cell(row=row, column=3, value=str(assessment.assessment_date))
            ws.cell(row=row, column=4, value=f"{assessment.risk_score:.1f}")
            ws.cell(row=row, column=5, value=assessment.get_risk_level_display())
            ws.cell(row=row, column=6, value=f"{assessment.ndvi_component_score:.1f}")
            ws.cell(row=row, column=7, value=f"{assessment.soil_moisture_component_score:.1f}")
            ws.cell(row=row, column=8, value=f"{assessment.weather_component_score:.1f}")
            ws.cell(row=row, column=9, value=f"{assessment.predicted_risk_7_days:.1f}" if assessment.predicted_risk_7_days else 'N/A')
            ws.cell(row=row, column=10, value=f"{assessment.predicted_risk_30_days:.1f}" if assessment.predicted_risk_30_days else 'N/A')
            ws.cell(row=row, column=11, value=f"{assessment.confidence_score:.2f}")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_alerts_sheet(self, wb):
        """Add alerts sheet"""
        ws = wb.create_sheet(title="Alerts")
        
        headers = [
            'Alert ID', 'Region', 'Title', 'Type', 'Severity', 'Priority',
            'Status', 'Created Date', 'Sent Date', 'Recipients', 'Successful', 'Failed'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        alerts = Alert.objects.filter(
            created_at__date__range=[self.start_date, self.end_date]
        ).select_related('region').order_by('-created_at')
        
        for row, alert in enumerate(alerts, 2):
            ws.cell(row=row, column=1, value=alert.alert_id)
            ws.cell(row=row, column=2, value=alert.region.name)
            ws.cell(row=row, column=3, value=alert.title)
            ws.cell(row=row, column=4, value=alert.get_template_display() if hasattr(alert, 'get_template_display') else 'N/A')
            ws.cell(row=row, column=5, value=alert.severity_level if hasattr(alert, 'severity_level') else 'N/A')
            ws.cell(row=row, column=6, value=alert.priority)
            ws.cell(row=row, column=7, value=alert.status)
            ws.cell(row=row, column=8, value=str(alert.created_at.date()))
            ws.cell(row=row, column=9, value=str(alert.sent_at.date()) if alert.sent_at else 'Not sent')
            ws.cell(row=row, column=10, value=alert.total_recipients)
            ws.cell(row=row, column=11, value=alert.successfully_sent)
            ws.cell(row=row, column=12, value=alert.failed_sends)
    
    def _add_weather_sheet(self, wb):
        """Add weather data sheet"""
        ws = wb.create_sheet(title="Weather Data")
        
        headers = [
            'Region', 'Date', 'Temperature (°C)', 'Humidity (%)', 'Rainfall (mm)',
            'Wind Speed (km/h)', 'Solar Radiation', 'Barometric Pressure'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        weather_data = WeatherData.objects.filter(
            timestamp__date__range=[self.start_date, self.end_date]
        ).select_related('region').order_by('-timestamp')[:1000]  # Limit to 1000 records
        
        for row, data in enumerate(weather_data, 2):
            ws.cell(row=row, column=1, value=data.region.name)
            ws.cell(row=row, column=2, value=str(data.timestamp.date()))
            ws.cell(row=row, column=3, value=f"{data.temperature:.1f}" if data.temperature else 'N/A')
            ws.cell(row=row, column=4, value=f"{data.humidity:.1f}" if data.humidity else 'N/A')
            ws.cell(row=row, column=5, value=f"{data.rainfall:.1f}" if data.rainfall else 'N/A')
            ws.cell(row=row, column=6, value=f"{data.wind_speed:.1f}" if data.wind_speed else 'N/A')
            ws.cell(row=row, column=7, value=f"{data.solar_radiation:.1f}" if data.solar_radiation else 'N/A')
            ws.cell(row=row, column=8, value=f"{data.barometric_pressure:.1f}" if data.barometric_pressure else 'N/A')
    
    def _add_farmers_sheet(self, wb):
        """Add farmers data sheet"""
        ws = wb.create_sheet(title="Farmers")
        
        headers = [
            'User ID', 'Name', 'Phone', 'Region', 'Farm Size (acres)',
            'Main Crops', 'Irrigation Type', 'Alert Preferences', 'Registration Date'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        farmers = FarmerProfile.objects.filter(
            user__date_joined__date__range=[self.start_date, self.end_date]
        ).select_related('user', 'region')
        
        for row, farmer in enumerate(farmers, 2):
            ws.cell(row=row, column=1, value=farmer.user.id)
            ws.cell(row=row, column=2, value=f"{farmer.user.first_name} {farmer.user.last_name}")
            ws.cell(row=row, column=3, value=farmer.phone_number or 'N/A')
            ws.cell(row=row, column=4, value=farmer.region.name if farmer.region else 'N/A')
            ws.cell(row=row, column=5, value=f"{farmer.farm_size:.1f}" if farmer.farm_size else 'N/A')
            ws.cell(row=row, column=6, value=farmer.main_crops or 'N/A')
            ws.cell(row=row, column=7, value=farmer.irrigation_type or 'N/A')
            ws.cell(row=row, column=8, value=farmer.alert_preferences or 'N/A')
            ws.cell(row=row, column=9, value=str(farmer.user.date_joined.date()))
    
    def _add_summary_pdf_content(self, story, styles):
        """Add summary content to PDF"""
        # System Overview
        story.append(Paragraph("System Overview", styles['Heading2']))
        
        # Statistics
        total_regions = Region.objects.filter(region_type='county').count()
        total_assessments = DroughtRiskAssessment.objects.filter(
            assessment_date__range=[self.start_date, self.end_date]
        ).count()
        total_alerts = Alert.objects.filter(
            created_at__date__range=[self.start_date, self.end_date]
        ).count()
        total_farmers = FarmerProfile.objects.count()
        
        overview_data = [
            ['Metric', 'Value'],
            ['Total Regions Monitored', str(total_regions)],
            ['Risk Assessments in Period', str(total_assessments)],
            ['Alerts Sent in Period', str(total_alerts)],
            ['Registered Farmers', str(total_farmers)],
        ]
        
        overview_table = Table(overview_data)
        overview_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(overview_table)
        story.append(Spacer(1, 20))
        
        # Regional Risk Summary
        story.append(Paragraph("Regional Risk Summary", styles['Heading2']))
        
        regional_data = [['Region', 'Latest Risk Score', 'Risk Level', 'Alerts Sent']]
        
        regions = Region.objects.filter(region_type='county')[:10]  # Limit to 10 for PDF
        for region in regions:
            latest_assessment = DroughtRiskAssessment.objects.filter(
                region=region,
                assessment_date__range=[self.start_date, self.end_date]
            ).order_by('-assessment_date').first()
            
            alerts_count = Alert.objects.filter(
                region=region,
                created_at__date__range=[self.start_date, self.end_date]
            ).count()
            
            if latest_assessment:
                regional_data.append([
                    region.name,
                    f"{latest_assessment.risk_score:.1f}",
                    latest_assessment.get_risk_level_display(),
                    str(alerts_count)
                ])
            else:
                regional_data.append([region.name, 'No data', 'No data', str(alerts_count)])
        
        regional_table = Table(regional_data)
        regional_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(regional_table)
    
    def _add_detailed_pdf_content(self, story, styles):
        """Add detailed content to PDF"""
        # Add summary content first
        self._add_summary_pdf_content(story, styles)
        
        # Add more detailed sections
        story.append(Spacer(1, 30))
        story.append(Paragraph("Detailed Alert Information", styles['Heading2']))
        
        alerts = Alert.objects.filter(
            created_at__date__range=[self.start_date, self.end_date]
        ).select_related('region').order_by('-created_at')[:20]  # Limit to 20 for PDF
        
        if alerts:
            alert_data = [['Alert ID', 'Region', 'Title', 'Status', 'Created Date']]
            for alert in alerts:
                alert_data.append([
                    alert.alert_id,
                    alert.region.name,
                    alert.title[:50] + '...' if len(alert.title) > 50 else alert.title,
                    alert.status,
                    str(alert.created_at.date())
                ])
            
            alert_table = Table(alert_data)
            alert_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(alert_table)
        else:
            story.append(Paragraph("No alerts found in the specified period.", styles['Normal']))
